#!/usr/bin/env python3
"""
Logbook Parser & Irradiation Run Assigner
==========================================
Reads campaign logbook Excel files to build run-number -> ion-species
mappings, then updates baselines_metadata.irrad_run_id for measurements
whose filenames contain _run{NNN}_.

Usage:
    python3 parse_logbooks_assign_runs.py [--dry-run]
"""

import argparse
import os
import re
import sys

import openpyxl
import psycopg2

from db_config import get_db_params, DATA_ROOT

IRRADIATION_ROOT = os.path.join(DATA_ROOT, "Measurements", "Irradiation")

# ── Logbook file paths (relative to IRRADIATION_ROOT) ──────────────────────
LOGBOOK_CONFIG = {
    "GSI_March_2025": os.path.join(
        "GSIMarch2025Au",
        "Copy of GSI_Au1162MeV_Mar2025_comments_NF_UG_NF.xlsx",
    ),
    "RADEF_2023": os.path.join(
        "21_RADEF Test Campaign 2023",
        "RADEF_June_2023",
        "LOGBOOK_21_06_2023.xlsx",
    ),
    "PSI_Proton_2022": os.path.join(
        "2022_30_08_PSI",
        "LOGBOOK_PSI_30_08_2022.xlsx",
    ),
    "ANSTO_Microbeam_2024": os.path.join(
        "ANSTO_23_01_2024_06_02_2024",
        "LOGBOOK_ANSTO_23_01_2024.xlsx",
    ),
    "UCL_Ions_2023": os.path.join(
        "2023_03_08_UCL_ions",
        "UCL_03_2023_analysis", "UCL",
        "UCL_Test_campaign_08_03_2023_DATA_Backup",
        "LOGBOOK_09_03_2023.xlsx",
    ),
    "GSI_Ca_2022": os.path.join(
        "2022_01_06_GSI_Ca",
        "Raw_data_GSI_all_currents", "Current_measurements",
        "2022_05_31_Ca",
        "LOGBOOK_2022_05_31.xlsx",
    ),
}


# ── Ion string normalisation ───────────────────────────────────────────────

def normalise_ion(raw):
    """Normalise logbook ion strings to match irradiation_runs.ion_species.

    Examples: 'Kr 769' -> 'Kr', 'C-36' -> 'C', 'Cl-36' -> 'Cl',
              'Ni-62' -> 'Ni', 'Ca' -> 'Ca'.
    """
    if not raw:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Strip isotope mass number after dash (C-36 -> C, Ni-62 -> Ni)
    s = re.sub(r'-\d+$', '', s)
    # Strip trailing numbers/spaces (Kr 769 -> Kr)
    s = re.sub(r'\s+\d+$', '', s)
    return s.strip() or None


def safe_int(val):
    """Try to parse a cell value as int, return None on failure."""
    if val is None:
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def safe_float(val):
    """Try to parse a cell value as float, return None on failure."""
    if val is None:
        return None
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


# ── Campaign-specific logbook parsers ──────────────────────────────────────
# Each returns {run_number: {"ion": str, ...extra fields}}

def parse_gsi_march2025(wb):
    """GSI March 2025 Au logbook.  All runs are Au (1162 MeV)."""
    mapping = {}
    sheet = wb["GSI Au "] if "GSI Au " in wb.sheetnames else wb["GSI Au"]
    for row in sheet.iter_rows(min_row=1, values_only=True):
        run_num = safe_int(row[0])
        if run_num is None:
            continue
        mapping[run_num] = {
            "ion": "Au",
            "board": safe_int(row[1]),
            "dut": safe_int(row[2]),
            "device": str(row[3]).strip() if row[3] else None,
        }
    return mapping


def parse_radef(wb):
    """RADEF 2023 logbook.  Multiple sheets, Ion column present."""
    mapping = {}
    # Sheet2 is the main run log
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        # Find header row containing 'Run' and 'Ion'
        header_idx = None
        ion_col = None
        run_col = None
        dut_col = None
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c else '' for c in row]
            if 'run' in cells and 'ion' in cells:
                header_idx = i
                run_col = cells.index('run')
                ion_col = cells.index('ion')
                dut_col = cells.index('dut') if 'dut' in cells else None
                break
        if header_idx is None:
            continue
        for row in rows[header_idx + 1:]:
            run_num = safe_int(row[run_col])
            if run_num is None:
                continue
            ion = normalise_ion(row[ion_col])
            if not ion:
                continue
            dut = str(row[dut_col]).strip() if dut_col is not None and row[dut_col] else None
            mapping[run_num] = {"ion": ion, "dut": dut}
    return mapping


def parse_psi(wb):
    """PSI Proton 2022.  Single ion (proton) for all runs."""
    mapping = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c else '' for c in row]
            if 'run' in cells or 'run ' in cells:
                run_col = next(
                    j for j, c in enumerate(cells) if c.startswith('run')
                )
                for data_row in rows[i + 1:]:
                    run_num = safe_int(data_row[run_col])
                    if run_num is not None:
                        mapping[run_num] = {"ion": "proton"}
                break
    return mapping


def parse_ansto(wb):
    """ANSTO Microbeam 2024.  Multi-sheet, has Ion + Energy columns."""
    mapping = {}
    # Only parse 'commercial' and date-named sheets (skip wafer sheets)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        # Find header row containing 'Run' in col 0 (or near it)
        header_idx = None
        run_col = ion_col = energy_col = boards_col = None
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c else '' for c in row]
            # Look for a row where one cell starts with "run"
            for j, c in enumerate(cells):
                if c.startswith('run'):
                    run_col = j
                    break
            else:
                continue
            # Found run column, now find ion and energy
            for j, c in enumerate(cells):
                if c == 'ion':
                    ion_col = j
                elif c.startswith('energy'):
                    energy_col = j
                elif c == 'boards':
                    boards_col = j
            if ion_col is not None:
                header_idx = i
                break
        if header_idx is None:
            continue
        for row in rows[header_idx + 1:]:
            run_num = safe_int(row[run_col])
            if run_num is None:
                continue
            ion = normalise_ion(row[ion_col])
            if not ion:
                continue
            energy = safe_float(row[energy_col]) if energy_col is not None else None
            board = str(row[boards_col]).strip() if boards_col is not None and row[boards_col] else None
            mapping[run_num] = {"ion": ion, "energy_mev": energy, "board": board}
    return mapping


def parse_ucl(wb):
    """UCL Ions 2023.  Multi-sheet per device type, Ion column present."""
    mapping = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        # Find header row
        header_idx = None
        run_col = ion_col = board_col = dut_col = None
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c else '' for c in row]
            if 'run' in cells and 'ion' in cells:
                header_idx = i
                run_col = cells.index('run')
                ion_col = cells.index('ion')
                board_col = cells.index('board') if 'board' in cells else None
                if board_col is None:
                    board_col = cells.index('board (wafer)') if 'board (wafer)' in cells else None
                dut_col = cells.index('dut') if 'dut' in cells else None
                break
        if header_idx is None:
            continue
        for row in rows[header_idx + 1:]:
            run_num = safe_int(row[run_col])
            if run_num is None:
                continue
            ion = normalise_ion(row[ion_col])
            if not ion:
                continue
            board = str(row[board_col]).strip() if board_col is not None and row[board_col] else None
            dut = str(row[dut_col]).strip() if dut_col is not None and row[dut_col] else None
            mapping[run_num] = {"ion": ion, "board": board, "dut": dut}
    return mapping


def parse_gsi_ca(wb):
    """GSI Ca 2022.  All runs are Ca."""
    mapping = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        header_idx = None
        run_col = dut_col = ion_col = None
        for i, row in enumerate(rows):
            cells = [str(c).strip().lower() if c else '' for c in row]
            if 'run' in cells:
                header_idx = i
                run_col = cells.index('run')
                dut_col = cells.index('dut') if 'dut' in cells else None
                ion_col = cells.index('ion') if 'ion' in cells else None
                break
        if header_idx is None:
            continue
        for row in rows[header_idx + 1:]:
            run_num = safe_int(row[run_col])
            if run_num is None:
                continue
            ion = normalise_ion(row[ion_col]) if ion_col is not None else "Ca"
            dut = str(row[dut_col]).strip() if dut_col is not None and row[dut_col] else None
            mapping[run_num] = {"ion": ion or "Ca", "dut": dut}
    return mapping


PARSERS = {
    "GSI_March_2025":        parse_gsi_march2025,
    "RADEF_2023":            parse_radef,
    "PSI_Proton_2022":       parse_psi,
    "ANSTO_Microbeam_2024":  parse_ansto,
    "UCL_Ions_2023":         parse_ucl,
    "GSI_Ca_2022":           parse_gsi_ca,
}


# ── Run lookup helpers ─────────────────────────────────────────────────────

def load_irrad_runs(cur, campaign_id):
    """Return {(ion_species, beam_energy_mev): run_id} for a campaign.

    For runs where beam_energy_mev is NULL, the key is (ion, None).
    """
    cur.execute(
        "SELECT id, ion_species, beam_energy_mev "
        "FROM irradiation_runs WHERE campaign_id = %s",
        (campaign_id,),
    )
    runs = {}
    for row in cur.fetchall():
        runs[(row[1], row[2])] = row[0]
    return runs


def find_run_id(runs_map, ion, energy_mev=None):
    """Find the irradiation_run id for a given ion and optional energy.

    For campaigns with a single run per ion (energy=NULL in DB), matches
    by ion alone.  For ANSTO where multiple energies exist per ion,
    matches on (ion, energy).
    """
    # Exact match first
    if (ion, energy_mev) in runs_map:
        return runs_map[(ion, energy_mev)]
    # Try ion with NULL energy (single-energy campaigns)
    if (ion, None) in runs_map:
        return runs_map[(ion, None)]
    # Try ion ignoring energy (if only one run for this ion)
    ion_matches = [(k, v) for k, v in runs_map.items() if k[0] == ion]
    if len(ion_matches) == 1:
        return ion_matches[0][1]
    return None


# ── Main assignment logic ──────────────────────────────────────────────────

def assign_runs_for_campaign(cur, campaign_id, campaign_name, logbook_map,
                             runs_map, dry_run=False):
    """Match measurements to irradiation runs via logbook run numbers.

    Returns (assigned, skipped_pre, unmatched_list).
    """
    # Fetch all measurements for this campaign that lack irrad_run_id
    cur.execute(
        "SELECT id, filename, device_id, irrad_role "
        "FROM baselines_metadata "
        "WHERE irrad_campaign_id = %s AND irrad_run_id IS NULL",
        (campaign_id,),
    )
    rows = cur.fetchall()

    assigned = 0
    skipped_pre = 0
    unmatched = []

    for meta_id, filename, device_id, irrad_role in rows:
        # Extract run number from filename (allow optional letter suffix: run031b)
        m = re.search(r'_run(\d+)[a-z]?[_.]', filename)
        if not m:
            # Files without run numbers (e.g. pre-characterisation CSVs)
            if irrad_role == 'pre_irrad':
                skipped_pre += 1
                continue
            unmatched.append((meta_id, filename, "no _run{NNN}_ in filename"))
            continue

        run_num = int(m.group(1))

        # run000 = pre-irrad characterisation
        if run_num == 0:
            skipped_pre += 1
            continue

        # Look up in logbook
        if run_num not in logbook_map:
            unmatched.append(
                (meta_id, filename, f"run {run_num} not in logbook")
            )
            continue

        entry = logbook_map[run_num]
        ion = entry["ion"]
        energy = entry.get("energy_mev")

        # Find matching irradiation_run
        run_id = find_run_id(runs_map, ion, energy)
        if run_id is None:
            unmatched.append(
                (meta_id, filename,
                 f"no irradiation_run for ion={ion} energy={energy}")
            )
            continue

        if not dry_run:
            cur.execute(
                "UPDATE baselines_metadata SET irrad_run_id = %s WHERE id = %s",
                (run_id, meta_id),
            )
        assigned += 1

    return assigned, skipped_pre, unmatched


def main():
    parser = argparse.ArgumentParser(
        description="Parse logbooks and assign irrad_run_id to measurements."
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview assignments without writing to the database.",
    )
    parser.add_argument(
        "--create-missing-runs", action="store_true",
        help="Auto-create irradiation_runs entries for ions found in "
             "logbooks but missing from the DB.",
    )
    args = parser.parse_args()

    conn = psycopg2.connect(**get_db_params())
    cur = conn.cursor()

    # Load campaign id mapping
    cur.execute("SELECT id, campaign_name FROM irradiation_campaigns")
    campaigns = {name: cid for cid, name in cur.fetchall()}

    print("=" * 72)
    print("Irradiation Run Assignment via Logbook Parsing")
    if args.dry_run:
        print("  ** DRY RUN — no database changes will be made **")
    print("=" * 72)
    print()

    summary = []

    for campaign_name, logbook_rel in LOGBOOK_CONFIG.items():
        print(f"--- {campaign_name} ---")

        if campaign_name not in campaigns:
            print(f"  SKIP: campaign not found in database\n")
            continue
        campaign_id = campaigns[campaign_name]

        logbook_path = os.path.join(IRRADIATION_ROOT, logbook_rel)
        if not os.path.isfile(logbook_path):
            print(f"  SKIP: logbook not found: {logbook_path}\n")
            continue

        # Count total unassigned
        cur.execute(
            "SELECT COUNT(*) FROM baselines_metadata "
            "WHERE irrad_campaign_id = %s AND irrad_run_id IS NULL",
            (campaign_id,),
        )
        total = cur.fetchone()[0]
        if total == 0:
            print(f"  No unassigned measurements.\n")
            summary.append((campaign_name, 0, 0, 0, 0))
            continue

        # Parse logbook
        parse_fn = PARSERS[campaign_name]
        wb = openpyxl.load_workbook(logbook_path, data_only=True)
        logbook_map = parse_fn(wb)
        wb.close()
        print(f"  Logbook entries parsed: {len(logbook_map)}")

        # Load irradiation_runs for this campaign
        runs_map = load_irrad_runs(cur, campaign_id)
        print(f"  Irradiation runs in DB: {len(runs_map)}")
        for (ion, energy), rid in sorted(runs_map.items()):
            print(f"    id={rid}: {ion} @ {energy} MeV")

        # Detect ions in logbook that have no irradiation_run entry
        logbook_ions = set()
        for entry in logbook_map.values():
            ion = entry["ion"]
            energy = entry.get("energy_mev")
            logbook_ions.add((ion, energy))
        missing_ions = [
            (ion, energy) for ion, energy in logbook_ions
            if find_run_id(runs_map, ion, energy) is None
        ]
        if missing_ions and args.create_missing_runs:
            for ion, energy in sorted(missing_ions):
                if args.dry_run:
                    # Simulate creation so assignment counts are accurate
                    fake_id = -(len(runs_map) + 1)
                    runs_map[(ion, energy)] = fake_id
                    print(f"  Would create run: {ion} @ {energy} MeV")
                else:
                    cur.execute(
                        "INSERT INTO irradiation_runs "
                        "(campaign_id, ion_species, beam_energy_mev) "
                        "VALUES (%s, %s, %s) RETURNING id",
                        (campaign_id, ion, energy),
                    )
                    new_id = cur.fetchone()[0]
                    runs_map[(ion, energy)] = new_id
                    conn.commit()
                    print(f"  Created run id={new_id}: {ion} @ {energy} MeV")
        elif missing_ions:
            print(f"  WARNING: logbook has ions not in DB "
                  f"(use --create-missing-runs to add):")
            for ion, energy in sorted(missing_ions):
                print(f"    {ion} @ {energy} MeV")

        # Assign
        assigned, skipped_pre, unmatched = assign_runs_for_campaign(
            cur, campaign_id, campaign_name, logbook_map,
            runs_map, dry_run=args.dry_run,
        )

        if not args.dry_run:
            conn.commit()

        print(f"  Total unassigned: {total}")
        print(f"  Assigned:         {assigned}")
        print(f"  Skipped (pre):    {skipped_pre}")
        print(f"  Unmatched:        {len(unmatched)}")
        if unmatched:
            for meta_id, fname, reason in unmatched[:10]:
                print(f"    [{meta_id}] {fname}: {reason}")
            if len(unmatched) > 10:
                print(f"    ... and {len(unmatched) - 10} more")
        print()

        summary.append((campaign_name, total, assigned, skipped_pre,
                         len(unmatched)))

    # Print summary table
    print("=" * 72)
    print(f"{'Campaign':<28s} {'Total':>6s} {'Assign':>6s} "
          f"{'Pre':>6s} {'Unmatched':>9s}")
    print("-" * 72)
    grand_total = grand_assigned = grand_pre = grand_unmatched = 0
    for name, total, asgn, pre, unmatch in summary:
        print(f"{name:<28s} {total:>6d} {asgn:>6d} {pre:>6d} {unmatch:>9d}")
        grand_total += total
        grand_assigned += asgn
        grand_pre += pre
        grand_unmatched += unmatch
    print("-" * 72)
    print(f"{'TOTAL':<28s} {grand_total:>6d} {grand_assigned:>6d} "
          f"{grand_pre:>6d} {grand_unmatched:>9d}")
    print("=" * 72)

    conn.close()


if __name__ == "__main__":
    main()
